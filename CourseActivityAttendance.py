todaydate=""
yesterdaydate=""
username=""
password=""
directory=""
#creates a dictionary of names and urls
#creates individual documents instead of a master sheet
#creates initial data structure from downloaded apex report
import openpyxl
wb= openpyxl.load_workbook('%s\\report1.xlsx' % directory)
sheet=wb.active
studentdict={}

for x in range(1,sheet.max_row):
    name= sheet.cell(row=x, column=19).value
    if name not in list(studentdict.keys()):
        studentdict[name]={}
    enrollmentid= sheet.cell(row=x, column=13).value
    subject= sheet.cell(row=x, column=9).value
    urllist=studentdict[name]
    urllist['https://reports-prd.apexvs.com/ApexUI/Reports/Student/courseActivityScoreReport.aspx?enrollmentID=%s' %(enrollmentid)]={'coursename' : subject, 'datecount': {}}
    



#selenium stuff
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os, time


#login and such
os.chdir(directory)
browser = webdriver.Firefox()
browser.get('http://apexvs.com')
elementz = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.NAME, 'ctl00$ContentPlaceHolder1$loginUsernameTextBox')))
login = browser.find_element_by_name('ctl00$ContentPlaceHolder1$loginUsernameTextBox')
login.send_keys(username)
passwd= browser.find_element_by_name('ctl00$ContentPlaceHolder1$passwordTextBox')
passwd.send_keys(password)
passwd.submit()
time.sleep(3) 

#scrapes the web data
studentlist=list(studentdict.keys())
#random.shuffle(studentlist)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = 'Name'
sheet.cell(row=1, column=2).value = 'Count'

for student in studentlist: #names
    #print(student)
    os.chdir(directory)
    a=studentdict[student]
    urllist= list(a.keys())
    todaycount=0
    yesterdaycount=0
    try:
        
        for y in urllist:  #urls
        
            #print(y)
            browser.get(y)
            cell = []
            celltext = []
            
            for x in range(100):   #populates the cell list with all completed assignments
                element = WebDriverWait(browser, 2).until(EC.presence_of_element_located((By.ID, 'reportGrid_cell_0_4')))
                cell.append(browser.find_elements_by_id('reportGrid_cell_%s_4' % x))
            for unit in cell:  #converts webelement to date
                for point in unit:
                    if len(point.text)>1:
                        celltext.append(point.text)
            
            
            b=a[y] #points to the value for the url key - should give coursename and datecount dict as output
            #print(b)
            c=b['datecount']
            
            for apexdate in celltext: #creates a dictionary with key=date and value= number of assignments on that date

                middate= apexdate.split() #added
                            
                if 'Jan' in apexdate:
                    middate[1]='01'
                elif 'Feb' in apexdate:
                    middate[1]='02'
                elif 'Mar' in apexdate:
                    middate[1]='03'
                elif 'Apr' in apexdate:
                    middate[1]='04'
                elif 'May' in apexdate:
                    middate[1]='05'
                elif 'Jun' in apexdate:
                    middate[1]='06'
                elif 'Jul' in apexdate:
                    middate[1]='07'
                elif 'Aug' in apexdate:
                    middate[1]='08'
                elif 'Sep' in apexdate:
                    middate[1]='09'
                elif 'Oct' in apexdate:
                    middate[1]='10'
                elif 'Nov' in apexdate:
                    middate[1]='11'
                elif 'Dec' in apexdate:
                    middate[1]='12'


                newdate= "%s/%s/%s" % (middate[1],middate[0],middate[2])
                if newdate not in list(c.keys()):
                    c[newdate]=1
                else:
                    c[newdate]+=1
            
            for d in list(c.keys()):
                if d==todaydate:
                    todaycount+=c[d]
                    print(todaycount)
                if d==yesterdaydate:
                    yesterdaycount+=c[d]
#deposits it in Excel

                
        
        
        
        
  #populate the worksheet
#        for name in list(studentdict.keys()):
        
    except Exception as err:
            pass    
    
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'Name'
    sheet.cell(row=1, column=2).value = 'Today Count'
    sheet.cell(row=1, column=3).value = 'Present or absent'
    sheet.cell(row=1, column=4).value = 'Final Count'
    
    a=studentdict[student]
    maxrow=sheet.max_row +1
  
    sheet.cell(row=maxrow, column=1).value = student
    sheet.cell(row=maxrow, column=2).value = todaycount
    sheet.cell(row=maxrow, column=3).value = ('=IF(B%s>0,\"P\",\"U\")' % maxrow)
    sheet.cell(row=maxrow, column=4).value = yesterdaycount
    
    
  
wb.save('%s\\attendance1.xlsx' % directory)  

